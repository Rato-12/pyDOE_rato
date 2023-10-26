# -*- coding: utf-8 -*-
"""
@author: Rato-12
"""

import io
import warnings
from pathlib import Path
from typing import Union

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import pyDOE2
import seaborn as sns

try:
    from design_of_experiment._def_screen import definitive_screening
except Exception:
    from _def_screen import definitive_screening

from matplotlib import rcParams

warnings.simplefilter('ignore', FutureWarning)
rcParams['font.family'] = 'sans-serif'
rcParams['font.sans-serif'] = ['Hiragino Maru Gothic Pro',
                               'Yu Gothic', 'Meirio', 'Takao',
                               'IPAGothic', 'IPAPGothic',
                               'VL PGothic', 'Noto Sans CJK JP']

sns.set_theme(style="whitegrid", context="notebook", font=['Yu Gothic'])


# 決定的スクリーニング計画
def defscreen(my_factors: dict, save: str = ""):
    dsd = definitive_screening(len(my_factors))
    # 3水準 [-1, 0, 1] を [0, 0.5, 1] に変換する
    dsd = dsd.where(dsd != 0.0, 0.5).where(dsd != -1.0, 0)
    print("Definitive_screening_design_sampling")
    fixed_data_np = range_to_value(my_factors, dsd.to_numpy().T)
    dsd_DF = pd.DataFrame(fixed_data_np.T)
    dsd_DF.columns = list(my_factors.keys())
    if save:
        dsd_DF.to_csv(
            f"{save}definitive screening.csv", sep=",",
            index=False, encoding="utf-8")
        plot_matrix_scatter(f"{save}definitive_screening",
                            dsd_DF, "winter")
    return dsd_DF, f"{save}definitive_screening"


# モンテカルロ法
def monte_carlo(my_factors: dict, sampling_num: int = 50, save: str = ""):
    monte_carlo_result = np.random.uniform(
        size=(len(my_factors), sampling_num))
    print("monte carlo_sampling")
    fixed_data_np = range_to_value(my_factors, monte_carlo_result)
    monte_carlo_DF = pd.DataFrame(fixed_data_np.T)
    monte_carlo_DF.columns = list(my_factors.keys())
    if save:
        monte_carlo_DF.to_csv(f"{save}monte_carlo.csv", sep=",",
                              index=False, encoding="utf-8")
        plot_matrix_scatter(f"{save}monte_carlo", monte_carlo_DF, "mako")
    return monte_carlo_DF, f"{save}monte_carlo"


# ラテン超方格法
def latin_hypercube(my_factors: dict, sampling_num: int = 50, save: str = ""):
    latin_hypercube_result = pyDOE2.lhs(len(my_factors), sampling_num).T
    print("latin_hypercube_sampling")
    fixed_data_np = range_to_value(my_factors, latin_hypercube_result)
    latin_hypercube_DF = pd.DataFrame(fixed_data_np.T)
    latin_hypercube_DF.columns = list(my_factors.keys())
    if save:
        latin_hypercube_DF.to_csv(
            f"{save}latin_hypercube.csv",
            sep=",", index=False, encoding="utf-8")
        plot_matrix_scatter(f"{save}latin_hypercube",
                            latin_hypercube_DF, "autumn")
    return latin_hypercube_DF, f"{save}latin_hypercube"


# 正規乱数
def normal_random(my_factors: dict, sampling_num: int = 50, save: str = ""):
    buf = []
    for cnt in range(len(my_factors)):
        buf.append(np.random.normal(0.5, 0.16, sampling_num))
    normal_random_result = np.array(buf)
    print("normal_random_sampling")
    fixed_data_np = range_to_value(my_factors, normal_random_result)
    normal_random_DF = pd.DataFrame(
        fixed_data_np.T, columns=list(my_factors.keys()))
    if save:
        normal_random_DF.to_csv(
            f"{save}normal_random.csv", sep=",", index=False, encoding="utf-8")
        plot_matrix_scatter(f"{save}normal_random", normal_random_DF, "gray")
    return normal_random_DF, f"{save}normal_random"


# ベータ乱数
def beta_random(my_factors: dict, sampling_num: int = 50, save: str = ""):
    buf = []
    for ent in range(len(my_factors)):
        buf.append(np.random.beta(5, 2, sampling_num))
    beta_random_result = np.array(buf)
    print("beta_random_sampling")
    fixed_data_np = range_to_value(my_factors, beta_random_result)
    beta_random_DF = pd.DataFrame(fixed_data_np.T,
                                  columns=list(my_factors.keys()))
    if save:
        beta_random_DF.to_csv(
            f"{save}beta_random.csv", sep=",",
            index=False, encoding="utf-8")
        plot_matrix_scatter(f"{save}beta_random", beta_random_DF, "spring_r")
    return beta_random_DF, f"{save}beta_random"


def d_optimal_select(df_data: pd.DataFrame,
                     sampling_num: int = 24, save: str = "",
                     exp_done_index_list: list = []) -> pd.DataFrame:
    """
    d最適計画法によって追加実験条件を計算
    Args:
        df data (pd.DataFrame) :
            実施済み+追加条件候補を記載したデータフレーム
        sampling num (int, optional):
            新たに追加する設定条件数 Defaults to 24.
        save (str, optional):
            結果の保存フォルダパス Defaults to "".
        exp_done_index_list (list, optional):
            実施済み条件のインデックスリスト。 df_data より選択 Defaults to [].
    Returns:
        pandas.DataFrame: d_最適化計画の結果。 実施済みも併記
    """

    print("D-optimal designing")
    number_of_selecting_samples = sampling_num  # 追加選択するサンプル数
    number_of_random_searches = 1000  # D最適基準を計算する繰り返し回数
    autoscaled_x = (df_data - df_data.mean()) / df_data.std()
    # 実験条件の候補のインデックスの作成
    all_indexes = list(range(df_data.shape[0]))
    exp_done_index = np.array(exp_done_index_list)
    # D最適基準に基づくサンプル選択
    np.random.seed(12)  # 乱数を生成するためのシードを固定
    for random_search_number in range(number_of_random_searches):
        # 1. ランダムに候補を選択
        new_selected_indexes = np.random.choice(
            all_indexes, number_of_selecting_samples, replace=False)

        searching_index = np.hstack((exp_done_index, new_selected_indexes))
        new_selected_samples = autoscaled_x.iloc[searching_index, :]

        # 2.D最適基準計算
        xt_x = np.dot(new_selected_samples.T, new_selected_samples)
        d_optimal_value = np.linalg.det(xt_x)
        # 3.D 最適基準が前回までの最大値を上回ったら、選択された候補を更
        if random_search_number == 0:
            best_d_optimal_value = d_optimal_value.copy()
            selected_sample_indexes = new_selected_indexes.copy()
        else:
            if best_d_optimal_value < d_optimal_value:
                best_d_optimal_value = d_optimal_value.copy()
                selected_sample_indexes = new_selected_indexes.copy()
        # 実施済みと新たに選択されたインデックスを統合
        done_next_index = np.hstack((exp_done_index, selected_sample_indexes))
        selected_samples = df_data.iloc[done_next_index, :]
        # print(selected_samples.corr())  # 相関行列
        selected_samples = selected_samples.reset_index(drop=True)
        selected_samples.to_csv(f"{save}d_optimal.csv", sep=",",
                                index=False, encoding="utf-8")
        plot_matrix_scatter(f"{save}d_optimal", selected_samples, "summer")
        return selected_samples, f"{save}d_optimal"


# 以下ユーティリティ
#
# 区間0~1の乱数データを指定した最小値~最大値の区間へ変換する関数


def range_to_value(my_factors, my_random_np_array):
    my_factors_column_names = list(my_factors.keys())
    for i, key in enumerate(my_factors_column_names):
        my_min, my_max = my_factors[key]
        my_random_np_array[i] = (
            (my_max - my_min) * my_random_np_array[i] + my_min)
    fixed_data_np = np.array(my_random_np_array)
    return fixed_data_np


# 行列散布図を作成する関数
def plot_matrix_scatter(label: str, DF: pd.DataFrame, my_color: str):
    sns.set(style="ticks", font_scale=1.2, palette=my_color, color_codes=True)
    g = sns.pairplot(DF, diag_kind="hist")
    g.fig.suptitle(label.split("//")[-1].split("\\")[-1]),  # fontsize=12)
    g.fig.subplots_adjust(top=0.9)
    plt.ioff()
    plt.savefig(label+'.png', bbox_inches="tight")
    plt.close()
    sns.reset_defaults()


#
#
def main(my_factors: dict, sampling_num: int = 24, save: str = ""):
    # ### Definitive_screening_design_sampling(DSD)
    df0 = defscreen(my_factors, save)
    print(df0)

    # ### Monte Carlo Sampling (MCS)
    df1 = monte_carlo(my_factors, sampling_num, save)
    print(df1)

    # ### Latin Hypercube Sampling (LHS)
    df2 = latin_hypercube(my_factors, sampling_num, save)
    print(df2)

    # ### Normal random Sampling
    df3 = normal_random(my_factors, sampling_num, save)
    print(df3)

    # Beta random Sampling
    df4 = beta_random(my_factors, sampling_num, save)
    print(df4)

    # ### Monte Carlo -> d-optimal
    df5 = monte_carlo(my_factors, 400)
    df5 = d_optimal_select(df5[0], sampling_num, save, [])
    print(df5)
    return df0, df1, df2, df3, df4, df5


# 処理の抽象化, Excelファイル入力、結果をExcelファイル出力
class ExpDesign():
    def __init__(self, xlsx_file: Union[str, pd.DataFrame],
                 sheet_name: str = 'condition',
                 save_path: str = None, method: str = None) -> None:
        """実験計画法による条件設計を行うクラス
        Args:
            xlsx_file (str): 読み込むExcelファイルのパス
            sheet_name (str, optional): 0. Defaults to 'condition'
            method (str, optional): 0. Defaults to None.
        """
        df_0 = None
        df_range = None
        my_factors = dict()

        if isinstance(xlsx_file, io.BytesIO):
            xlsx_file = xlsx_file
            suffix = '.xlsx'  # 事前にチェック済みだと思うので
        else:
            xlsx_file = Path(xlsx_file)
            suffix = xlsx_file.suffix  # ファイルパスの時は拡張子判定

        if suffix.lower() == '.xlsx':
            try:
                df_0 = pd.read_excel(
                    xlsx_file, header=0, index_col=0, sheet_name=sheet_name)
            except Exception:
                df_0 = pd.read_csv(
                    xlsx_file, skiprows=1, index_col=0,
                    engine="python", encoding="shift-jis")
                df_0 = df_0.dropna(how="all").dropna(axis='columns')
            df_range = df_0.copy().head(2)

            if method == "d_optimal_select":
                df_0 = df_0.copy().drop(
                    index=['Min', 'Max']).reset_index(drop=True)
                df_0 = df_0.dropna()  # 欠損行は対象外
            df_range = df_range.dropna(axis='columns')  # 欠損列は対象外
            if isinstance(df_range, pd.DataFrame):
                my_factors = df_range.to_dict(orient='list')
        # 保存先設定
        if save_path:
            savepath_str = f'{save_path}\\exp_design\\'
            savepath = Path(savepath_str)
            savepath.mkdir(parents=True, exist_ok=True)
        else:
            savepath_str = ""

        self.method = method  # 実験計画法の手法
        self.my_factors = my_factors  # 条件設計の対象変数とその範囲 辞書形式
        self.search_range = df_range  # 条件設計の対象変数とその範囲  pandas.DataFrame
        self.pre_condition = df_0  # 実施済みの条件 D基準最適化の時のみ有効
        self.savepath_str = savepath_str  # 結果の保存先
        self.doe_df = ""  # 結果保存用の初期化
        self.pairplot_fig_path = ""  # 結果保存用の初期化
        self.xlsx_name = ""  # 結果保存用の初期化

    def run(self, sampling_num: int):
        method = self.method
        my_factors = self.my_factors
        save = self.savepath_str
        pre_data = self.pre_condition

        if method == "defscreen":
            # ### Definitive screening design sampling(DSD)
            df, fig = defscreen(my_factors, save)
            mess = "Definitive screening design sampling(DSD)"
        elif method == "monte_carlo":
            # ### Monte Carlo Sampling (MCS)
            df, fig = monte_carlo(my_factors, sampling_num, save)
            mess = "Monte Carlo Sampling (MCS)"
        elif method == "latin_hypercube":
            # ### Latin Hypercube Sampling (LHS)
            df, fig = latin_hypercube(my_factors, sampling_num, save)
            mess = "Latin Hypercube Sampling (LHS)"
        elif method == "d_optimal_select":
            # ### Monte Carlo -> d-optimal
            df, fig = monte_carlo(my_factors, 500)
            mess = "Monte Carlo Sampling -> D-Optimal Selection"
            pre_index = []
            if isinstance(pre_data, pd.DataFrame):
                if list(pre_data.columns) == list(df.columns):
                    pre_index = list(pre_data.reset_index(drop=True).index)
                    df = pd.concat([pre_data, df], ignore_index=True)
                    sampling_num = sampling_num + len(pre_index) - 2
                else:
                    mess = "設定項目が一致しないため、 実施済み条件は無視されました。"
            df, fig = d_optimal_select(df, sampling_num, save, pre_index)
        self.doe_df = df
        self.pairplot_fig_path = fig
        return df, mess, fig

    def to_excel(self, start_row: int, start_col: int):
        # Excelへ出力
        import datetime

        import openpyxl
        from openpyxl.styles.borders import Border, Side

        doe_df = self.doe_df
        df_range = self.search_range
        save_path = self.savepath_str

        wb = openpyxl.Workbook()
        data_sheet = wb.create_sheet(title="condition", index=0)

        # DataFrame書き出し
        if list(doe_df.columns) == list(df_range.columns):
            export_df = pd.concat([df_range, doe_df])
        else:
            export_df = doe_df
        for x in range(0, len(export_df.index)):
            for y in range(0, len(export_df.columns)):
                data_sheet.cell(
                    x + start_row + 1, y+start_col + 1
                ).value = export_df.iat[x, y]
        for x, name in enumerate(list(export_df.index)):
            data_sheet.cell(x + start_row + 1, start_col).value = name
        for y, name in enumerate(list(export_df.columns)):
            data_sheet.cell(start_row, y + start_col + 1).value = name

        # 書式設定反映
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side)
        for y, _ in enumerate(list(export_df.columns)):
            data_sheet.cell(start_row, y + start_col + 1).border = border
        border = Border(right=side)
        for x, _ in enumerate(list(export_df.index)):
            data_sheet.cell(x + start_row + 1, start_col).border = border
        border = Border(top=side, bottom=side, right=side)
        data_sheet.cell(start_row, start_col). border = border

        # 図形貼り付け
        graph_sheet = wb.create_sheet(title="pairplot", index=1)
        img = openpyxl.drawing.image.Image(f'{self.pairplot_fig_path}.png')
        img.width = 600
        img.height = 600
        graph_sheet.add_image(img, 'A1')

        # Excel 保存 + 閉じる
        _date = datetime.datetime.now().strftime('%Y%m%d%H%M%S%f_')
        xlsx_name = Path(f'{save_path}{_date}DOE.xlsx')
        wb.save(xlsx_name)
        wb.close()
        self.xlsx_name = xlsx_name
        return xlsx_name


if __name__ == "__main__":
    import os

    sampling_num = 48  # 生成数設定
    # 因子名とその乱数を生成する上下限を辞書で設定
    my_factors = {
        "height": (50, 200),
        "width": (0.06, 0.1),
        "density": (1e15, 9e15),
        "temp": (-50, 250),
    }

    savepath_str = f"{os.getcwd()}\\exp_design_result\\"
    savepath = Path(savepath_str)
    savepath.mkdir(parents=True, exist_ok=True)

    main(my_factors, sampling_num, savepath_str)
